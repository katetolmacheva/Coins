import os
import threading
import asyncio
from flask import Flask, render_template, request, jsonify, send_from_directory
import telegram
from telegram import Update, WebAppInfo, ReplyKeyboardMarkup, KeyboardButton
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes
from sqlalchemy import create_engine, Column, Integer, String, Float
from sqlalchemy.orm import sessionmaker, declarative_base
import openpyxl
from openpyxl.utils import get_column_letter
import platform
import asyncio
import logging
import datetime
import re 

logging.getLogger("urllib3").setLevel(logging.WARNING)
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# --- Конфигурация ---
TOKEN = "7790860611:AAEJ7y8BJlSRkNhva6Uaxdg04vjIpCU-sbE" # ЗАМЕНИТЬ НА СВОЙ ТОКЕН БОТА ИЗ BOTFATHER!
EXCEL_FILE = 'coins.xlsx'
ORDERS_EXCEL_FILE = 'orders.xlsx'
DATABASE_FILE = 'database.db'
DATABASE_URL = f'sqlite:///{DATABASE_FILE}'

# ЕСЛИ ДЕПЛОЙИТ НА СЕРВЕР, ЗАМЕНИТЬ ЭТОТ URL НА ПУБЛИЧНЫЙ АДРЕС СЕРВЕРА!
FLASK_APP_BASE_URL = 'https://980aa8da38c1.ngrok-free.app' 
ADMIN_USER_ID = [1043419485 
                 #,123456789, 
                 #987654321
                 ] # ЗАМЕНИТЬ НА ID ПОЛЬЗОВАТЕЛЕЙ, КОТОРЫЕ БУДЕТ ИМЕТЬ ПРАВА АДМИНИСТРАТОРА

OFFICES = [
    "Барвиха" , "Денежный", "Ордынка", "Пресненский", "Цифровой", "Казанская", "Басков переулок", "Владивосток", "Казань", "Нижний Новгород", "Краснодар", "Красноярск", "Екатеринбург", "Новосибирск", "Самара", "Ростов-на-Дону", "Тюмень", "Челябинск"
]

app = Flask(__name__, static_folder='static', template_folder='templates')

# --- Настройка SQLAlchemy ---
Base = declarative_base()

class Coin(Base):
    __tablename__ = 'coins'
    id = Column(Integer, primary_key=True)
    name = Column(String, nullable=False)
    number = Column(String, unique=True, nullable=False)
    available_quantity = Column(Integer, default=0)
    denomination = Column(String)
    material = Column(String)
    price = Column(Float)
    file_name = Column(String)

    def __repr__(self):
        return f"<Coin(name='{self.name}', number='{self.number}', quantity={self.available_quantity})>"

engine = create_engine(DATABASE_URL)
Session = sessionmaker(bind=engine)

def init_db():
    Base.metadata.create_all(engine)
    logging.info("База данных инициализирована.")

REQUIRED_EXCEL_COLUMNS = [
    'Название',
    'Номер',
    'Доступное количество',
    'Номинал',
    'Металл',
    'Цена',
    'Имя файла'
]

def sync_excel_to_db(update: Update = None, context: ContextTypes.DEFAULT_TYPE = None) -> None:
    session = Session();
    if update:
        user_id = update.effective_user.id
        if user_id not in ADMIN_USER_ID:
            logging.warning(f"Попытка несанкционированного вызова sync_excel_to_db. User ID: {user_id}")
            asyncio.run(update.message.reply_text("Отказано в доступе. Обновление может выполнять только администратор."))
            return  

    logging.info("Начинаю синхронизацию данных из Excel...")
    try:
        if not os.path.exists(EXCEL_FILE):
            logging.error(f"Файл '{EXCEL_FILE}' не найден.")
            if update:
                asyncio.run(update.message.reply_text(f"Ошибка: файл '{EXCEL_FILE}' не найден. Обратитесь к администратору."))
            return
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        sheet = workbook.active
        header = [cell.value for cell in sheet[1]]

        for col in REQUIRED_EXCEL_COLUMNS:
            if col not in header:
                raise ValueError(f"Не найдена обязательная колонка в Excel-файле: {col}")

        session.query(Coin).delete()
        session.commit()
        logging.info("Существующие данные в базе данных очищены.")

        rows_processed = 0
        for row_index in range(2, sheet.max_row + 1):
            row_data = {header[i]: sheet.cell(row=row_index, column=i+1).value for i in range(len(header))}
            
            if not any(row_data.values()): 
                continue 

            if not row_data.get('Название'):
                logging.warning(f"Пропущена строка {row_index} из-за отсутствия 'Названия'.")
                continue

            try:
                available_quantity = int(row_data.get('Доступное количество', 0) or 0)
            except (ValueError, TypeError):
                available_quantity = 0
                logging.warning(f"Некорректное значение 'Доступное количество' в строке {row_index}: {row_data.get('Доступное количество')}. Установлено 0.")

            try:
                price = float(str(row_data.get('Цена', 0.0)).replace(',', '.') or 0.0)
            except (ValueError, TypeError):
                price = 0.0
                logging.warning(f"Некорректное значение 'Цена' в строке {row_index}: {row_data.get('Цена')}. Установлено 0.0.")

            coin = Coin(
                name=row_data.get('Название'),
                number=row_data.get('Номер'),
                available_quantity=available_quantity,
                denomination=row_data.get('Номинал'),
                material=row_data.get('Металл'),
                price=price,
                file_name=row_data.get('Имя файла')
            )
            session.add(coin)
            rows_processed += 1
            
        session.commit()
        logging.info(f"Данные успешно синхронизированы из Excel в базу данных. Обработано {rows_processed} строк.")
        return True 

    except ValueError as e:
        logging.error(f"Ошибка при синхронизации данных из Excel: {e}")
        session.rollback()
        return False
    except Exception as e:
        logging.exception(f"Произошла непредвиденная ошибка при синхронизации Excel:")
        session.rollback()
        return False
    finally:
        session.close()

def clear_orders_excel():
    """Удаляет существующий файл orders.xlsx, чтобы он создался заново."""
    if os.path.exists(ORDERS_EXCEL_FILE):
        try:
            os.remove(ORDERS_EXCEL_FILE)
            logging.info(f"Существующий файл заказов '{ORDERS_EXCEL_FILE}' удален.")
        except Exception as e:
            logging.error(f"Ошибка при удалении файла заказов '{ORDERS_EXCEL_FILE}': {e}")
    else:
        logging.info(f"Файл заказов '{ORDERS_EXCEL_FILE}' не найден, создавать его не нужно.")


def save_order_to_excel(order_details):
    try:
        if not os.path.exists(ORDERS_EXCEL_FILE):
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Заказы"
            sheet.append([
                "Дата и время", 
                "Имя", 
                "Фамилия", 
                "Офис", 
                "Название монеты", 
                "Код монеты", 
                "Количество"
            ])
        else:
            workbook = openpyxl.load_workbook(ORDERS_EXCEL_FILE)
            sheet = workbook["Заказы"]
        
        for item in order_details:
            sheet.append(item)
        
        workbook.save(ORDERS_EXCEL_FILE)
        logging.info(f"Заказ успешно сохранен в файл Excel: {ORDERS_EXCEL_FILE}")
    except Exception as e:
        logging.error(f"Ошибка при сохранении заказа в Excel: {e}")



@app.route('/')
@app.route('/index.html')
def index():
    return render_template('index.html')

@app.route('/cart.html')
def cart():
    return render_template('cart.html')

@app.route('/static/<path:filename>')
def static_files(filename):
    return send_from_directory(app.static_folder, filename)

@app.route('/api/coins', methods=['GET'])
def get_coins():
    session = Session()
    try:
        query = session.query(Coin)

        search_query = request.args.get('search', '').lower()
        sort_by = request.args.get('sort_by', 'name')
        sort_order = request.args.get('sort_order', 'asc')
        material_filter = request.args.get('material', '')
        denomination_filter = request.args.get('denomination', '')
        availability_filter = request.args.get('availability', '')

        if search_query:
            query = query.filter(
                (Coin.name.ilike(f'%{search_query}%')) |
                (Coin.number.ilike(f'%{search_query}%'))
            )

        if material_filter:
            query = query.filter(Coin.material == material_filter)

        if denomination_filter:
            query = query.filter(Coin.denomination == denomination_filter)

        if availability_filter:
            if availability_filter == 'in_stock':
                query = query.filter(Coin.available_quantity > 0)
            elif availability_filter == 'out_of_stock':
                query = query.filter(Coin.available_quantity <= 0)

        if sort_by == 'name':
            if sort_order == 'desc':
                query = query.order_by(Coin.name.desc())
            else:
                query = query.order_by(Coin.name.asc())
        elif sort_by == 'price':
            if sort_order == 'desc':
                query = query.order_by(Coin.price.desc())
            else:
                query = query.order_by(Coin.price.asc())

        coins = query.all()

        coins_data = [
            {
                'name': c.name,
                'number': c.number,
                'available_quantity': c.available_quantity,
                'denomination': c.denomination,
                'material': c.material,
                'price': c.price,
                'file_name': c.file_name
            }
            for c in coins
        ]
        return jsonify(coins_data)

    except Exception as e:
        logging.error(f"Ошибка при получении монет: {e}")
        return jsonify({'error': 'Ошибка сервера при получении данных'}), 500
    finally:
        session.close()

@app.route('/api/reserve/<string:coin_number>', methods=['POST'])
def reserve_coin(coin_number):
    session = Session()
    try:
        coin = session.query(Coin).filter_by(number=coin_number).first()
        if coin and coin.available_quantity > 0:
            coin.available_quantity -= 1
            session.commit()
            return jsonify({'success': True, 'new_quantity': coin.available_quantity})
        elif coin and coin.available_quantity == 0:
            return jsonify({'success': False, 'message': 'Монета закончилась.'})
        else:
            return jsonify({'success': False, 'message': 'Монета не найдена.'})
    except Exception as e:
        session.rollback()
        logging.error(f"Ошибка бронирования монеты {coin_number}: {e}")
        return jsonify({'success': False, 'message': f'Ошибка бронирования: {str(e)}'})
    finally:
        session.close()

@app.route('/api/order', methods=['POST'])
def handle_order():
    session = Session()
    try:
        data = request.json
        logging.info(f"Получен запрос на заказ. Raw JSON Data: {data}")

        customer_info = data.get('customer_info')
        cart_items = data.get('cart_items')

        if not customer_info or not cart_items:
            logging.warning(f"Отсутствуют данные о клиенте или товарах в корзине. customer_info: {customer_info}, cart_items: {cart_items}")
            return jsonify({'success': False, 'message': 'Отсутствуют данные о клиенте или товарах в корзине.'}), 400

        logging.info(f"Информация о клиенте: {customer_info}")
        logging.info(f"Товары в корзине: {cart_items}")

        results = []
        all_items_available = True

        for item in cart_items:
            coin_number = item.get('number')
            requested_quantity = item.get('quantity')
            
            logging.info(f"Обработка позиции: coin_number={coin_number}, requested_quantity={requested_quantity}")

            if coin_number is None or requested_quantity is None:
                logging.warning(f"Пропущена позиция из корзины: отсутствует 'number' или 'quantity'. Item: {item}")
                continue

            try:
                requested_quantity = int(requested_quantity)
            except (ValueError, TypeError):
                logging.error(f"Некорректное значение quantity для монеты {coin_number}: {requested_quantity}")
                all_items_available = False
                results.append({'coin_number': coin_number, 'status': 'invalid_quantity', 'requested_quantity': item.get('quantity')})
                continue

            coin = session.query(Coin).filter_by(number=coin_number).first()

            if coin:
                logging.info(f"Найдена монета '{coin.name}' (Номер: {coin.number}). Доступно: {coin.available_quantity}. Запрошено: {requested_quantity}.")
                if coin.available_quantity >= requested_quantity:
                    coin.available_quantity -= requested_quantity
                    results.append({'coin_number': coin_number, 'status': 'success', 'quantity_booked': requested_quantity})
                    logging.info(f"Монета {coin.number}: забронировано {requested_quantity}. Остаток: {coin.available_quantity}")
                else:
                    all_items_available = False
                    status = 'not_enough_stock'
                    available = coin.available_quantity
                    results.append({'coin_number': coin_number, 'status': status, 'requested_quantity': requested_quantity, 'available_quantity': available})
                    logging.warning(f"Монета {coin_number}: Недостаточно на складе. Запрошено {requested_quantity}, доступно {available}. Оформление заказа отменено для этой монеты.")
            else:
                all_items_available = False
                status = 'not_found'
                results.append({'coin_number': coin_number, 'status': status, 'requested_quantity': requested_quantity, 'available_quantity': 0})
                logging.warning(f"Монета {coin_number}: Не найдена в базе данных. Оформление заказа отменено для этой монеты.")

        if all_items_available:
            logging.info("Все позиции доступны. Пытаюсь зафиксировать изменения в базе данных.")
            session.commit()
            logging.info("Все монеты успешно забронированы. Сохраняю заказ в Excel.")

            order_details = []
            order_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            for item in cart_items:
                coin = session.query(Coin).filter_by(number=item.get('number')).first() 
                order_details.append([
                    order_time,
                    customer_info.get('name', ''),
                    customer_info.get('surname', ''),
                    customer_info.get('office', ''),
                    coin.name if coin else "Неизвестно", 
                    item.get('number'),
                    item.get('quantity')
                ])
            save_order_to_excel(order_details)

            logging.info("Заказ успешно оформлен и сохранен.")
            return jsonify({'success': True, 'message': 'Заказ успешно оформлен!', 'details': results})
        else:
            session.rollback()
            logging.warning("Оформление заказа отменено из-за нехватки монет или других проблем. Все изменения в БД отменены.")
            return jsonify({'success': False, 'message': 'Не удалось оформить весь заказ. Проверьте детали.', 'details': results})
            
    except Exception as e:
        session.rollback()
        logging.exception(f"Произошла ошибка при оформлении заказа:")
        return jsonify({'success': False, 'message': f'Произошла ошибка при оформлении заказа: {str(e)}', 'details': []})
    finally:
        session.close()

@app.route('/api/checkout', methods=['POST'])
def checkout():
    data = request.get_json()
    items_to_reserve = data.get('items', [])
    user_first_name = data.get('user_first_name', 'Не указано')
    user_last_name = data.get('user_last_name', 'Не указано')
    user_office = data.get('user_office', 'Не указано')

    session = Session()

    results = []
    success_all = True
    
    try:
        for item in items_to_reserve:
            coin_number = item.get('number')
            quantity = item.get('quantity', 1)
            coin = session.query(Coin).filter_by(number=str(coin_number)).first()
            
            if not coin:
                results.append({'number': coin_number, 'success': False, 'message': 'Монета не найдена.'})
                success_all = False
                break
            
            if coin.available_quantity < quantity:
                results.append({
                    'number': coin_number, 
                    'success': False, 
                    'message': f'Недостаточно монет в наличии для "{coin.name}". Доступно: {coin.available_quantity}, Запрошено: {quantity}.'
                })
                success_all = False
                break
        
        if success_all:
            order_records_for_excel = []
            current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            order_items_text = []

            for item in items_to_reserve:
                coin_number = item.get('number')
                quantity = item.get('quantity', 1)
                coin = session.query(Coin).filter_by(number=str(coin_number)).first()
                if coin:
                    coin.available_quantity -= quantity
                    session.add(coin)

                    order_records_for_excel.append([
                        current_time,
                        user_first_name,
                        user_last_name,
                        user_office,
                        coin.name,
                        coin.number,
                        quantity
                    ])
                    order_items_text.append(f"• Название: {coin.name}\n  Номер: {coin.number}\n  Количество: {quantity}\n")

            session.commit()
            logging.info(f"Заказ от '{user_first_name} {user_last_name}' успешно зарезервирован в БД.")

            save_order_to_excel(order_records_for_excel)

            message_items_str = "\n".join(order_items_text)
            message = f"Новый заказ!\n\nИмя: {user_first_name}\nФамилия: {user_last_name}\nОфис: {user_office}\n\nТовары в заказе:\n{message_items_str}"

            if telegram_application_instance and telegram_bot_loop and telegram_bot_loop.is_running():
                for admin_id in ADMIN_USER_ID:
                    try:
                        # Корректный способ отправить сообщение через asyncio
                        future = asyncio.run_coroutine_threadsafe(
                            telegram_application_instance.bot.send_message(chat_id=admin_id, text=message),
                            telegram_bot_loop
                        )
                        future.result(timeout=10)
                        logging.info(f"Уведомление успешно отправлено админу {admin_id}.")
                    except Exception as e:
                        logging.error(f"Не удалось отправить уведомление админу {admin_id}: {e}")
            else:
                logging.error("Не удалось отправить уведомление: Telegram бот не запущен или цикл событий недоступен.")

            return jsonify({'success': True, 'message': 'Заказ успешно оформлен!', 'details': results})

              
        else:
            session.rollback()
            logging.warning("Оформление заказа отменено из-за нехватки монет.")
            return jsonify({'success': False, 'message': 'Не удалось оформить весь заказ. Проверьте детали.', 'details': results})
            
    except Exception as e:
        session.rollback()
        logging.exception(f"Произошла ошибка при оформлении заказа:")
        return jsonify({'success': False, 'message': f'Произошла ошибка при оформлении заказа: {str(e)}', 'details': []})
    finally:
        session.close()


# --- Функции Telegram-бота ---

telegram_application_instance: Application = None
telegram_bot_loop: asyncio.AbstractEventLoop = None

telegram_application_instance = None
telegram_bot_loop = None

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Обрабатывает команду /start."""
    # Используем заранее определенный FLASK_APP_BASE_URL
    webapp_url = f"{FLASK_APP_BASE_URL}/index.html"
    keyboard = [[KeyboardButton("Открыть каталог", web_app=WebAppInfo(url=webapp_url))]]
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True, one_time_keyboard=False)

    await update.message.reply_text(
        "Привет! Я бот для заказа монет. Нажмите 'Открыть каталог', чтобы начать.",
        reply_markup=reply_markup
    )

async def update_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Обработчик команды /update_excel_data. Обновляет данные из Excel файла."""
    user_id = update.effective_user.id
    logging.info(f"Получена команда /update_excel_data от пользователя с ID: {user_id}")
    
    if user_id in ADMIN_USER_ID:
        message_text = "Начинаю обновление базы данных из Excel файла..."
        await update.message.reply_text(message_text)
        threading.Thread(target=sync_excel_to_db, args=(update, context)).start()
        message_text = "Каталог обновлен!"
        await update.message.reply_text(message_text)
    else:
        logging.warning(f"Попытка обновить данные не-администратором. User ID: {user_id}")
        await update.message.reply_text("У вас нет прав для выполнения этой команды.")

def run_flask_app():
    app.run(host='0.0.0.0', port=5000, debug=False, use_reloader=False)

def run_telegram_bot_in_thread():
    global telegram_application_instance, telegram_bot_loop

    telegram_bot_loop = asyncio.new_event_loop()
    asyncio.set_event_loop(telegram_bot_loop)

    telegram_application_instance = Application.builder().token(TOKEN).build()

    telegram_application_instance.add_handler(CommandHandler("start", start))
    telegram_application_instance.add_handler(CommandHandler("update", update_command)) 

    logging.info("Запуск Telegram бота (polling) в отдельном потоке...")
    try:
        # Инициализация бота перед запуском polling
        telegram_bot_loop.run_until_complete(telegram_application_instance.initialize())
        logging.info("Telegram бот инициализирован.")

        telegram_bot_loop.run_until_complete(
            telegram_application_instance.run_polling(allowed_updates=Update.ALL_TYPES, drop_pending_updates=True, timeout=30)
        )
    except telegram.error.TimedOut:
        logging.error("Фатальная ошибка в потоке Telegram бота: Timed out. Проверьте соединение с серверами Telegram или настройки прокси.")
    except Exception as e:
        logging.error(f"Фатальная ошибка в потоке Telegram бота: {e}", exc_info=True)
    finally:
        logging.info("Telegram бот остановлен.")
        if telegram_application_instance.running: # Проверяем, запущен ли бот
            telegram_bot_loop.call_soon_threadsafe(telegram_application_instance.shutdown)
        if telegram_bot_loop and not telegram_bot_loop.is_closed():
            telegram_bot_loop.close()

if __name__ == '__main__':
    if platform.system() == "Windows":
        asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())
    
    clear_orders_excel()
    logging.info("Файл заказов orders.xlsx подготовлен к новой сессии.")

    init_db()
    sync_excel_to_db()

    flask_thread = threading.Thread(target=run_flask_app, daemon=True)
    flask_thread.start()
    logging.info("Flask-приложение запущено в отдельном потоке на порту 5000...")

    telegram_thread = threading.Thread(target=run_telegram_bot_in_thread, daemon=True)
    telegram_thread.start()
    logging.info("Поток Telegram бота запущен...")

    try:
        while True:
            import time
            time.sleep(1)
    except KeyboardInterrupt:
        logging.info("Приложение остановлено пользователем (Ctrl+C).")
        if telegram_application_instance:
            logging.info("Останавливаю Telegram бота...")
            telegram_bot_loop.call_soon_threadsafe(telegram_application_instance.stop)
            telegram_thread.join(timeout=5)
            if telegram_thread.is_alive():
                logging.warning("Поток Telegram бота не завершился вовремя.")
    except Exception as e:
        logging.exception(f"Произошла непредвиденная ошибка в основном потоке:")


def run_flask_app():
    app.run(host='0.0.0.0', port=5000, debug=False, use_reloader=False)

if __name__ == '__main__':
    if platform.system() == "Windows":
        asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())
    
    clear_orders_excel()
    logging.info("Файл заказов orders.xlsx подготовлен к новой сессии.")

    init_db()
    sync_excel_to_db()

    flask_thread = threading.Thread(target=run_flask_app, daemon=True)
    flask_thread.start()
    logging.info("Flask-приложение запущено в отдельном потоке на порту 5000...")

    try:
        while True:
            import time
            time.sleep(1)
    except KeyboardInterrupt:
        logging.info("Приложение остановлено пользователем (Ctrl+C).")
    except Exception as e:
        logging.exception(f"Произошла непредвиденная ошибка в основном потоке:")